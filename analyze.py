import glob
import json
import os
import re
import time
from datetime import date as date_type
from urllib.parse import parse_qs, urlparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from groq import Groq
from dotenv import load_dotenv

load_dotenv()

# Fallback order: best quality first, lighter models as backup
GROQ_MODELS = [
    "llama-3.3-70b-versatile",
    "llama-3.1-8b-instant",
    "gemma2-9b-it",
]

_groq_client = None


def _get_client() -> Groq:
    global _groq_client
    if _groq_client is None:
        key = os.environ.get("GROQ_API_KEY")
        if not key:
            raise SystemExit("Missing GROQ_API_KEY in .env")
        _groq_client = Groq(api_key=key)
    return _groq_client


def _call_groq(prompt: str) -> str:
    client = _get_client()
    last_error = None
    for model in GROQ_MODELS:
        for attempt in range(3):
            try:
                response = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=600,
                    temperature=0.3,
                )
                text = response.choices[0].message.content
                if model != GROQ_MODELS[0]:
                    print(f"    (used fallback model: {model})")
                return text
            except Exception as e:
                msg = str(e)
                last_error = e
                if "rate_limit" in msg.lower() or "429" in msg:
                    delay_match = re.search(r"try again in ([\d.]+)s", msg)
                    delay = float(delay_match.group(1)) + 2 if delay_match else 20
                    print(f"    Rate limited on {model} — waiting {delay:.0f}s (attempt {attempt + 1}/3)...")
                    time.sleep(delay)
                else:
                    # Non-rate-limit error (e.g. decommissioned) — skip to next model immediately
                    print(f"    {model} unavailable: skipping to next model...")
                    break
    raise last_error


def load_latest_reviews(data_dir: str = "data") -> list[dict]:
    files = glob.glob(os.path.join(data_dir, "reviews_*.json"))
    if not files:
        raise SystemExit(f"No review files found in {data_dir}/. Run scrape.py first.")
    latest = max(files, key=os.path.getmtime)
    print(f"Loading reviews from {latest}")
    with open(latest) as f:
        return json.load(f)


def _extract_query(url: str) -> str:
    qs = parse_qs(urlparse(url).query)
    return qs.get("query", [""])[0].lower().strip()


def _addr_key(address: str) -> tuple[str, str] | None:
    """Extract (street_number, zip_code) as a unique store identifier."""
    number = re.match(r"^(\d+)", address.strip())
    zip_code = re.search(r"\b(\d{5})\b", address)
    if number and zip_code:
        return (number.group(1), zip_code.group(1))
    return None


def group_reviews_by_store(reviews: list[dict], stores: list[dict]) -> list[dict]:
    store_by_query = {_extract_query(s["google_maps_url"]): s for s in stores}
    # Address key: (street_number, zip) — most reliable since it comes from Google Maps data
    store_by_addr_key: dict[tuple, dict] = {}
    for s in stores:
        key = _addr_key(s["address"])
        if key:
            store_by_addr_key[key] = s

    grouped: dict[str, dict] = {s["google_maps_url"]: {"store": s, "reviews": []} for s in stores}

    for review in reviews:
        # Priority 1: match by review's address field (Google Maps location data — most accurate)
        store = None
        rev_key = _addr_key(review.get("address", ""))
        if rev_key:
            store = store_by_addr_key.get(rev_key)

        # Priority 2: searchString match
        if store is None:
            search_str = review.get("searchString", "").lower().strip()
            store = store_by_query.get(search_str) if search_str else None

        # Priority 3: URL query match
        if store is None:
            query = _extract_query(review.get("url", ""))
            store = store_by_query.get(query)

        if store:
            grouped[store["google_maps_url"]]["reviews"].append(review)

    return list(grouped.values())


def calculate_avg_rating(reviews: list[dict]) -> float | None:
    ratings = [r["stars"] for r in reviews if r.get("stars") is not None]
    if not ratings:
        return None
    return round(sum(ratings) / len(ratings), 1)


def build_gemini_prompt(store_name: str, address: str, reviews: list[dict]) -> str:
    review_lines = []
    for i, r in enumerate(reviews, 1):
        stars = r.get("stars", "?")
        text = (r.get("text") or "").strip()
        if text:
            review_lines.append(f"{i}. [{stars} stars] {text}")

    reviews_block = "\n".join(review_lines) if review_lines else "No review text available."

    return f"""You are an operations analyst for Yum and Chill Restaurant Group, a fast-food franchise operator.

Analyze the following {len(reviews)} Google Maps customer reviews for {store_name} located at {address}.

REVIEWS:
{reviews_block}

Provide a concise analysis for the store manager in this exact format:

SUMMARY:
• [1-2 sentences describing a specific recurring theme, referencing what customers actually said]
• [1-2 sentences describing a specific recurring theme, referencing what customers actually said]
• [1-2 sentences describing a specific recurring theme, referencing what customers actually said]

IMPROVEMENTS:
• [2-5 word improvement area]
• [2-5 word improvement area]
• [2-5 word improvement area]

For SUMMARY: be specific — mention actual patterns (e.g. wait times, specific menu items, staff behavior). Reference the volume of complaints where relevant.
For IMPROVEMENTS: 2-5 words only, no sentences."""


def parse_gemini_response(raw: str) -> dict:
    summary_bullets: list[str] = []
    improvement_bullets: list[str] = []
    section = None

    for line in raw.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if line.upper().startswith("SUMMARY"):
            section = "summary"
        elif line.upper().startswith("IMPROVEMENTS") or line.upper().startswith("IMPROVEMENT"):
            section = "improvements"
        elif line.startswith("•") or line.startswith("-"):
            text = line.lstrip("•-").strip()
            if text:
                if section == "summary":
                    summary_bullets.append(text)
                elif section == "improvements":
                    improvement_bullets.append(text)

    summary = "\n".join(f"• {b}" for b in summary_bullets) if summary_bullets else raw.strip()[:500]
    return {"summary": summary, "actions": improvement_bullets}


def analyze_store(store_name: str, address: str, reviews: list[dict]) -> dict:
    if not reviews:
        return {
            "summary": "No reviews found in this period.",
            "actions": [],
            "avg_rating": None,
            "review_count": 0,
        }

    avg_rating = calculate_avg_rating(reviews)
    prompt = build_gemini_prompt(store_name, address, reviews)

    try:
        raw = _call_groq(prompt)
        parsed = parse_gemini_response(raw)
    except Exception as e:
        print(f"  WARNING: All models failed for {store_name}: {e}")
        parsed = {"summary": "Analysis unavailable.", "actions": []}

    return {
        "summary": parsed["summary"],
        "actions": parsed["actions"],
        "avg_rating": avg_rating,
        "review_count": len(reviews),
    }


BRAND_STYLES = {
    "Taco Bell": {"bg": "702082", "fg": "F5D619"},
    "Wendy's": {"bg": "CC2222", "fg": "FFFFFF"},
}

RATING_FILLS = {
    "red":    PatternFill("solid", fgColor="FFCCCC"),
    "yellow": PatternFill("solid", fgColor="FFFFCC"),
    "green":  PatternFill("solid", fgColor="CCFFCC"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

COLUMN_WIDTHS = {
    "Store #": 12,
    "Store Name": 28,
    "Full Address": 40,
    "Avg Rating": 12,
    "Reviews Analyzed": 18,
    "Period": 16,
    "AI Summary & Action Items": 80,
    "Customer Reviews": 70,
}
HEADERS = list(COLUMN_WIDTHS.keys())


def _format_reviews(reviews: list[dict]) -> str:
    lines = []
    for i, r in enumerate(reviews, 1):
        stars = r.get("stars", "?")
        text = (r.get("text") or "").strip()
        if text:
            if len(text) > 300:
                text = text[:297] + "..."
            lines.append(f"{i}. [{stars}★] {text}")
    return "\n".join(lines) if lines else "No review text available."


def _rating_fill(rating: float | None) -> PatternFill | None:
    if rating is None:
        return None
    if rating <= 3.4:
        return RATING_FILLS["red"]
    if rating <= 3.9:
        return RATING_FILLS["yellow"]
    return RATING_FILLS["green"]


def _write_sheet(ws, group_stores: list[dict], brand: str, period_label: str):
    style = BRAND_STYLES.get(brand, {"bg": "333333", "fg": "FFFFFF"})
    header_fill = PatternFill("solid", fgColor=style["bg"])
    header_font = Font(bold=True, color=style["fg"], size=11)

    # Write header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # Write data rows
    for row_idx, entry in enumerate(group_stores, 2):
        store = entry["store"]
        analysis = entry["analysis"]

        improvements_text = ""
        if analysis["actions"]:
            improvements_text = "IMPROVEMENTS:\n" + "\n".join(f"• {a}" for a in analysis["actions"])

        summary_full = "SUMMARY:\n" + analysis["summary"]
        if improvements_text:
            summary_full = f"{summary_full}\n\n{improvements_text}"

        rating = analysis["avg_rating"]
        rating_display = f"{rating:.1f} ⭐" if rating is not None else "—"

        reviews_text = _format_reviews(entry.get("reviews", []))

        row_values = [
            store["store_number"],
            store["name"],
            store["address"],
            rating_display,
            analysis["review_count"] or "—",
            period_label,
            summary_full,
            reviews_text,
        ]

        # Alternating row background
        row_fill = PatternFill("solid", fgColor="F9F9F9") if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx >= len(HEADERS) - 1))

            if row_fill:
                cell.fill = row_fill

            # Rating column conditional color (overrides row fill)
            if col_idx == 4:
                fill = _rating_fill(rating)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="top")

        combined_lines = summary_full.count("\n") + reviews_text.count("\n") + 6
        ws.row_dimensions[row_idx].height = max(80, 15 * combined_lines)

    # Set column widths
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[header]

    # Freeze header row
    ws.freeze_panes = "A2"


def write_excel(grouped: list[dict], start_date: str, out_dir: str = "output") -> str:
    os.makedirs(out_dir, exist_ok=True)
    month_label = date_type.fromisoformat(start_date).strftime("%b%Y")
    today_label = date_type.today().strftime("%b %d, %Y")
    period_label = f"{date_type.fromisoformat(start_date).strftime('%b %Y')} – {date_type.today().strftime('%b %d, %Y')}"
    filename = f"Y&C_Google_Review_Analysis_{month_label}.xlsx"
    out_path = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    tab_groups = [
        ("Taco Bell", "Taco Bell"),
        ("Wendy's North", "Wendy's"),
        ("Wendy's South", "Wendy's"),
    ]

    for tab_name, brand in tab_groups:
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_properties.tabColor = (
            "702082" if brand == "Taco Bell" else "CC2222"
        )
        group_stores = [e for e in grouped if e["store"]["group"] == tab_name]
        _write_sheet(ws, group_stores, brand, period_label)

    wb.save(out_path)
    print(f"Excel report saved to {out_path}")
    return out_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Analyze reviews and write Excel report")
    parser.add_argument("--start", required=True, help="Start date YYYY-MM-DD matching your scrape")
    args = parser.parse_args()

    with open("stores.json") as f:
        stores = json.load(f)

    reviews = load_latest_reviews()
    grouped_raw = group_reviews_by_store(reviews, stores)

    print(f"Analyzing {len(grouped_raw)} stores with Groq ({GROQ_MODELS[0]})...")
    grouped = []
    for entry in grouped_raw:
        store = entry["store"]
        store_reviews = entry["reviews"]
        print(f"  {store['name']} ({len(store_reviews)} reviews)")
        analysis = analyze_store(store["name"], store["address"], store_reviews)
        grouped.append({"store": store, "analysis": analysis, "reviews": store_reviews})
        if store_reviews:
            time.sleep(2)  # stay under Groq free tier RPM limit

    out_path = write_excel(grouped, args.start)
    print(f"\nDone. Report: {out_path}")


if __name__ == "__main__":
    main()
